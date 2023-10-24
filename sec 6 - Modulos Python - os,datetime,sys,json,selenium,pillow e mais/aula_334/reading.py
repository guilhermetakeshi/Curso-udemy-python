# openpyxl - Ler e alterar dados de uma planilha do Excel
# Com a biblioteca openpyxl, é possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos, criar fórmulas, adicionar
# imagens e outros elementos gráficos às planilhas do Excel. Ela é útil para
# automatizar tarefas envolvendo planilhas do Excel, como a criação de
# relatórios, análise de dados e a manipulação de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# Obtém o caminho para o diretório onde este script está localizado
ROOT_FOLDER = Path(__file__).parent

# Define o caminho para o arquivo de trabalho (workbook) Excel
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carrega um arquivo Excel existente
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome da planilha a ser selecionada
sheet_name = 'Minha planilha'

# Seleciona a planilha pelo nome
worksheet: Worksheet = workbook[sheet_name]

# Itera pelas linhas da planilha (começa a partir da segunda linha)
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        # Exemplo: Se o valor da célula for 'Maria', altera o valor na coluna B
        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23)

    print()

# Exemplo de como alterar diretamente o valor de uma célula específica
# worksheet['B3'].value = 14

# Salva as alterações de volta no arquivo Excel
workbook.save(WORKBOOK_PATH)

# openpyxl - Criando uma planilha do Excel (Workbook e Worksheet)
# Com a biblioteca openpyxl, é possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos, criar fórmulas, adicionar
# imagens e outros elementos gráficos às planilhas do Excel. Ela é útil para
# automatizar tarefas envolvendo planilhas do Excel, como a criação de
# relatórios, análise de dados e a manipulação de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Obtém o caminho para o diretório onde este script está localizado
ROOT_FOLDER = Path(__file__).parent

# Define o caminho para o arquivo de trabalho (workbook) Excel
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Cria uma instância de Workbook (uma pasta de trabalho)
workbook = Workbook()

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos na primeira linha da planilha
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

# Dados dos estudantes em uma lista
students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

# Popula a planilha com os dados dos estudantes
for student in students:
    worksheet.append(student)

# Salva a pasta de trabalho no caminho especificado
workbook.save(WORKBOOK_PATH)
